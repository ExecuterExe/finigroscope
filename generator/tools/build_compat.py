# -*- coding: utf-8 -*-
"""Собирает compat-data.js из листа «Совместимость параметров».

Запуск (из корня проекта):
    python tools/build_compat.py "C:/путь/к/Таблицы.xlsx"

Без аргумента берётся файл, указанный в DEFAULT_XLSX.

Лист устроен так: строки 4..127 и столбцы D..DW — одни и те же 124 параметра
(вариант ответа на вопрос опросника). Шапка столбцов: строка 1 — блок,
строка 2 — вопрос, строка 3 — вариант ответа. Ячейка на пересечении содержит
'+', '+/-', '-' либо пуста.

Подписи вариантов берём из шапки столбцов (строка 3): в столбце C первая
строка не заполнена, а строка 3 заполнена целиком. Оси симметричны, поэтому
подписи столбцов годятся и для строк — скрипт это проверяет.
"""

import sys
import io
import json
import datetime
from pathlib import Path

import openpyxl

DEFAULT_XLSX = r"C:/Users/Eugene/Downloads/Tablitsy_Sovmestimosti_I_Oprosnik_30_07_26.xlsx"
SHEET = "Совместимость параметров"
FIRST_ROW = 4
FIRST_COL = 4

# как значение ячейки кодируется в compat-data.js
CODE = {"+": "+", "+/-": "~", "-": "-", "": "."}

OUT = Path(__file__).resolve().parent.parent / "compat-data.js"


def merged_lookup(ws):
    """Значение объединённой ячейки видно только в её левом верхнем углу —
    растягиваем его на всю область, иначе шапка блоков/вопросов будет дырявой."""
    filled = {}
    for rng in ws.merged_cells.ranges:
        value = ws.cell(rng.min_row, rng.min_col).value
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                filled[(r, c)] = value
    return lambda r, c: filled.get((r, c), ws.cell(r, c).value)


def text(value):
    return "" if value is None else str(value).strip()


def qnum(question):
    """«4*.Виды ОВЗ» -> «4*», «10.Есть ли случайность...» -> «10»"""
    return text(question).split(".")[0]


def main():
    src = Path(sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX)
    if not src.is_file():
        sys.exit("Не найден файл: %s" % src)

    wb = openpyxl.load_workbook(src, data_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit("В книге нет листа «%s». Есть: %s" % (SHEET, wb.sheetnames))
    ws = wb[SHEET]
    val = merged_lookup(ws)

    n = ws.max_row - FIRST_ROW + 1
    if n != ws.max_column - FIRST_COL + 1:
        sys.exit("Матрица не квадратная: %d строк, %d столбцов"
                 % (n, ws.max_column - FIRST_COL + 1))

    axis = []
    for i in range(n):
        col = FIRST_COL + i
        axis.append({
            "block": text(val(1, col)),
            "q": qnum(val(2, col)),
            "question": text(val(2, col)),
            "opt": text(val(3, col)),
        })

    problems = []

    # оси должны совпадать: иначе ячейку нельзя читать «строка × столбец»
    for i in range(n):
        row = FIRST_ROW + i
        row_q, row_opt = qnum(val(row, 2)), text(val(row, 3))
        if row_q != axis[i]["q"]:
            problems.append("строка %d: вопрос «%s», а в столбце «%s»"
                            % (row, row_q, axis[i]["q"]))
        # подпись варианта в строке 4 не заполнена в исходнике — это известно
        if row_opt and row_opt != axis[i]["opt"]:
            problems.append("строка %d: вариант «%s», а в столбце «%s»"
                            % (row, row_opt, axis[i]["opt"]))

    grid = []
    unknown = set()
    for i in range(n):
        chars = []
        for j in range(n):
            raw = text(ws.cell(FIRST_ROW + i, FIRST_COL + j).value)
            if raw not in CODE:
                unknown.add(raw)
                raw = ""
            chars.append(CODE[raw])
        grid.append("".join(chars))

    if unknown:
        problems.append("неизвестные значения ячеек: %s" % sorted(unknown))

    # матрица обязана быть симметричной — иначе результат проверки зависел бы
    # от того, в каком порядке пользователь отвечал на вопросы
    for i in range(n):
        for j in range(i + 1, n):
            if grid[i][j] != grid[j][i]:
                problems.append("несимметрично: %s×%s = «%s», обратно «%s»"
                                % (axis[i]["opt"], axis[j]["opt"],
                                   grid[i][j], grid[j][i]))

    if problems:
        print("Проблемы в исходной таблице (%d):" % len(problems))
        for p in problems[:40]:
            print("  -", p)
        sys.exit(1)

    counts = {c: sum(row.count(c) for row in grid) for c in "+~-."}
    header = (
        "/* ФАЙЛ СОБРАН АВТОМАТИЧЕСКИ — правьте не его, а исходную таблицу.\n"
        "   Пересборка: python tools/build_compat.py \"путь/к/файлу.xlsx\"\n\n"
        "   Источник: %s\n"
        "   Лист: %s\n"
        "   Собрано: %s\n"
        "   Параметров: %d, пар проверено: %d\n"
        "   Ячейки: «+» %d, «+/-» %d, «-» %d, без данных %d\n"
        "   ========================================================= */\n\n"
        % (src.name, SHEET, datetime.date.today().isoformat(), n,
           n * (n - 1) // 2, counts["+"], counts["~"], counts["-"], counts["."])
    )

    lines = [header]
    lines.append("const COMPAT_DATA = {\n")
    lines.append("    source: %s,\n" % json.dumps(src.name, ensure_ascii=False))
    lines.append("    sheet: %s,\n" % json.dumps(SHEET, ensure_ascii=False))
    lines.append("    generated: %s,\n" % json.dumps(datetime.date.today().isoformat()))
    lines.append("\n    /* Параметры в том же порядке, что строки и столбцы листа.\n")
    lines.append("       q — номер вопроса опросника, opt — вариант ответа. */\n")
    lines.append("    axis: [\n")
    for a in axis:
        lines.append("        { q: %s, opt: %s },\n" % (
            json.dumps(a["q"], ensure_ascii=False),
            json.dumps(a["opt"], ensure_ascii=False)))
    lines.append("    ],\n")
    lines.append("\n    /* Строка i — совместимость параметра i со всеми остальными:\n")
    lines.append("       '+' совместимы, '~' («+/-») сочетание спорное,\n")
    lines.append("       '-' несовместимы, '.' в таблице не заполнено. */\n")
    lines.append("    grid: [\n")
    for i, row in enumerate(grid):
        lines.append("        %s%s  // %d %s\n" % (
            json.dumps(row), "," if i < n - 1 else "", i, axis[i]["opt"]))
    lines.append("    ]\n")
    lines.append("};\n")

    with io.open(OUT, "w", encoding="utf-8", newline="\n") as f:
        f.write("".join(lines))

    print("Готово: %s" % OUT)
    print("  параметров: %d, пар: %d" % (n, n * (n - 1) // 2))
    print("  «+» %d, «+/-» %d, «-» %d, без данных %d"
          % (counts["+"], counts["~"], counts["-"], counts["."]))


if __name__ == "__main__":
    main()
