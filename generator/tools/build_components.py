# -*- coding: utf-8 -*-
"""Excel «Компоненты» -> agents/library/components.json.

Три листа книги задают весь расчёт этапа 5, и все три нужны целиком:

  «Компоненты (опт. знач.)» — базовый диапазон и НИЖНИЙ ПРЕДЕЛ по компоненту;
  «Расчет диапазонов»       — шаг поправки на каждое значение каждого параметра;
  «Компоненты , материал»   — какие материалы допустимы для компонента.

Почему преобразователь, а не чтение xlsx в рантайме: сервер намеренно живёт на
стандартной библиотеке, а openpyxl — зависимость только для разработки. Заодно
таблица перестаёт быть непрозрачным двоичным файлом: расхождение видно в диффе.

Запуск (из generator/):
    python tools/build_components.py "путь/к/Tablitsy...xlsx"

Скрипт ГРОМКО жалуется на всё, что не сошлось с опросником: молчаливо
пропущенное значение означало бы шаг, который никогда не применится, и поймать
это потом можно только по неверным числам в готовой игре.
"""

import json
import re
import sys
from collections import OrderedDict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import params as params_module                                    # noqa: E402

BASE_DIR = Path(__file__).resolve().parent.parent
OUT_FILE = BASE_DIR / "agents" / "library" / "components.json"

SHEET_BASE = "Компоненты (опт. знач.)"
SHEET_RANGES = "Расчет диапазонов"
SHEET_MATERIAL = "Компоненты , материал"

# Названия компонентов в таблице -> канонические имена опросника (params.COMPONENTS).
# Таблица пишет их вольно («карты», «Карты», «Фигурки/миниатюры»), опросник —
# строго, и связывать их надо явно: опечатка здесь тихо обнулит весь расчёт.
COMPONENT_ALIASES = {
    "карты": "карты",
    "игровое поле": "игровое поле",
    "кубики": "кубики",
    "кубики (d6, d10, d20 и др.)": "кубики",
    "фишки": "фишки",
    "фишки (для игроков)": "фишки",
    "жетоны": "жетоны",
    "жетоны (для ресурсов/очков)": "жетоны",
    "таймер": "таймер",
    "песочные часы/таймер": "таймер",
    "песочные часы": "таймер",
    "фигурки": "фигурки",
    "фигурки/миниатюры": "фигурки",
    "фигурки/миниатюры (для самой игры)": "фигурки",
    "телефоны": "телефоны",
    "телефоны/смартфоны игроков": "телефоны",
}

# Названия параметров в таблице -> ключи game_params.
PARAM_ALIASES = {
    "возраст": "age_group",
    "жанр": "genre",
    "случайность": "randomness",
    "место игры": "location",
    "количество игроков": "player_count",
    "среднее время партии": "play_time",
    "сложность": "complexity",
    "адаптация": "adaptation",
    "адаптация для лиц с овз": "adaptation",
}

# Значения, означающие «любое» — шаг применяется всегда.
ANY_VALUES = {"любой", "любая", "любое", "любые"}

# Случайность и адаптация в game_params — булевы, а в таблице словесные.
RANDOMNESS_VALUES = {
    "есть случайность": True,
    "нет случайности": False,
    "полная детерминированность": False,
}
YES_NO = {"да": True, "нет": False}

DASHES = "–—−"          # –, —, − — в таблице встречаются все три


def norm(text):
    """Приводит строку к сравнимому виду: тире, пробелы, косые, регистр.

    Пробелы вокруг косой убираются намеренно: в книге одно и то же пишут и
    «фигурки/миниатюры», и «фигурки / миниатюры», и это не разные вещи.
    """
    if text is None:
        return ""
    text = str(text).strip().lower()
    for dash in DASHES:
        text = text.replace(dash, "-")
    text = re.sub(r"\s*/\s*", "/", text)
    return re.sub(r"\s+", " ", text)


# Значения параметров в книге записаны короче, чем в опроснике. Это те же
# самые варианты, поэтому связываем их явной таблицей, а не подгонкой norm():
# опечатка в такой подгонке склеила бы два РАЗНЫХ значения и осталась незаметной.
VALUE_ALIASES = {
    "genre": {
        "экономика/торговля": "экономика",
        "игра со словами": "слова и ассоциации",
        "игра на реакцию/память": "реакция и память",
        "игра с блефом": "блеф",
        "детектив/расследование": "детектив",
        "строительство и развитие": "строительство",
    },
    "player_count": {
        "10+ игроков": "10+ игроков (для больших компаний)",
    },
    "complexity": {
        "низкая (для новичков и детей)": "низкая",
        "средняя (для подготовленных игроков)": "средняя",
        "высокая (для опытных стратегов)": "высокая",
    },
}


def alias_value(param, value):
    return VALUE_ALIASES.get(param, {}).get(value, value)


# Префикс числовой корзины. Отличать её от варианта опросника нужно ЯВНО:
# иначе «2 игрока» (вариант) и «2-4 игрока» (корзина) лежали бы в одном
# пространстве имён и однажды перекрыли бы друг друга.
BUCKET_PREFIX = "#"


def parse_bucket(param, value):
    """«1-4 игрока» -> «#1-4», «9+ игроков» -> «#9-». None, если не корзина."""
    if param != "player_count":
        return None
    raw = value.replace(" ", "")
    match = re.match(r"^(\d+)-(\d+)игрок", raw)
    if match:
        return "%s%s-%s" % (BUCKET_PREFIX, match.group(1), match.group(2))
    match = re.match(r"^(\d+)\+игрок", raw)
    if match:
        return "%s%s-" % (BUCKET_PREFIX, match.group(1))
    return None


def parse_range(text):
    """«30–40 шт.» -> (30, 40); «1» -> (1, 1). None, если чисел нет."""
    numbers = re.findall(r"-?\d+", norm(text).replace("-", " -"))
    numbers = [int(n) for n in re.findall(r"\d+", norm(text))]
    if not numbers:
        return None
    if len(numbers) == 1:
        return (numbers[0], numbers[0])
    return (numbers[0], numbers[1])


def parse_step(text):
    """«–15» -> -15, «10» -> 10, «0» -> 0. None, если не число."""
    raw = norm(text)
    if not raw:
        return None
    match = re.match(r"^(-?)\s*(\d+)$", raw.replace(" ", ""))
    if not match:
        return None
    value = int(match.group(2))
    return -value if match.group(1) == "-" else value


class Report:
    """Копит расхождения, чтобы показать их разом и с числами."""

    def __init__(self):
        self.problems = []
        self.notes = []

    def problem(self, text):
        self.problems.append(text)

    def note(self, text):
        self.notes.append(text)

    def show(self):
        for line in self.notes:
            print("  " + line)
        if self.problems:
            print("\nРАСХОЖДЕНИЯ С ОПРОСНИКОМ (%d):" % len(self.problems))
            for line in self.problems:
                print("  ! " + line)
        return not self.problems


def component_key(raw, report):
    key = COMPONENT_ALIASES.get(norm(raw))
    if key is None:
        report.problem("неизвестный компонент: «%s»" % raw)
    return key


def read_base(wb, report):
    """Базовые диапазоны и нижние пределы."""
    ws = wb[SHEET_BASE]
    out = OrderedDict()
    for row in ws.iter_rows(values_only=True):
        cells = [c for c in row]
        if not cells or cells[0] is None:
            continue
        name = norm(cells[0])
        if name in ("компонент", "") or name.startswith("таблица"):
            continue
        if name.startswith("для "):
            continue
        if name == "другое":
            continue                      # индивидуально, формулой не считается
        key = component_key(cells[0], report)
        if key is None:
            continue
        base = parse_range(cells[1] if len(cells) > 1 else None)
        floor = parse_range(cells[2] if len(cells) > 2 else None)
        if base is None:
            report.problem("нет базового диапазона у «%s»" % cells[0])
            continue
        out[key] = {
            "base_min": base[0],
            "base_max": base[1],
            # Нижний предел — не рекомендация, а запрет: ниже него игра
            # перестаёт работать, сколько бы минусов ни набрали поправки.
            "floor": floor[0] if floor else base[0],
            "per_player": "на игрока" in norm(cells[1]),
            "why": str(cells[3]).strip() if len(cells) > 3 and cells[3] else "",
        }
    return out


def read_materials(wb, report):
    """Матрица «компонент x материал» со значениями +, +/-, -."""
    ws = wb[SHEET_MATERIAL]
    rows = list(ws.iter_rows(values_only=True))
    header = None
    out = OrderedDict()
    for row in rows:
        cells = list(row)
        if not cells or all(c is None for c in cells):
            continue
        first = norm(cells[0])
        if first.startswith("компоненты"):
            header = [str(c).strip() if c else "" for c in cells[1:]]
            continue
        if header is None or cells[0] is None:
            continue
        if first == "другое":
            continue
        key = component_key(cells[0], report)
        if key is None:
            continue
        allowed = OrderedDict()
        for name, value in zip(header, cells[1:]):
            if not name:
                continue
            allowed[name] = str(value).strip() if value else "-"
        out[key] = allowed
    return out


def _survey_values(param):
    """Допустимые значения параметра — прямо из опросника, а не из головы."""
    if param == "age_group":
        return set(norm(k) for k in params_module.AGE_RANGES)
    if param == "player_count":
        return set(norm(k) for k in params_module.PLAYER_RANGES)
    if param == "play_time":
        return set(norm(k) for k in params_module.PLAY_TIME_RANGES)
    if param == "genre":
        return set(norm(v) for v in params_module.GENRES.values())
    if param == "complexity":
        return set(norm(v) for v in params_module.COMPLEXITY.values())
    if param == "location":
        return None                # свободный список, сверять не с чем
    return None


def read_steps(wb, report):
    """Шаги поправки по каждому компоненту и параметру."""
    ws = wb[SHEET_RANGES]
    out = OrderedDict()
    component = None
    param = None
    seen_values = {}

    for row in ws.iter_rows(values_only=True):
        cells = list(row) + [None] * (7 - len(row))
        first = norm(cells[0])

        # Заголовок блока пишут и «Компонент: карты», и «Компоненты: другое».
        if first.startswith("компонент:") or first.startswith("компоненты:"):
            raw = str(cells[0]).split(":", 1)[1]
            if norm(raw) in ("другое", "другой"):
                # «Другое» считается индивидуально: формулы для него нет и быть
                # не может — что это за компонент, знает только автор.
                component, param = None, None
                continue
            component = component_key(raw, report)
            param = None
            if component and component not in out:
                out[component] = OrderedDict()
            continue

        if component is None or first == "параметр":
            continue

        if cells[0]:
            param = PARAM_ALIASES.get(first)
            if param is None:
                report.problem("неизвестный параметр «%s» у компонента «%s»"
                               % (cells[0], component))
                continue
            out[component].setdefault(param, OrderedDict())

        if param is None or not cells[1]:
            continue

        step = parse_step(cells[3])
        if step is None:
            continue

        value = alias_value(param, norm(cells[1]))
        entry = {"step": step,
                 "why": str(cells[5]).strip() if cells[5] else "",
                 "note": str(cells[6]).strip() if cells[6] else ""}

        if value in ANY_VALUES:
            out[component][param]["*"] = entry
            continue

        if param == "randomness":
            flag = RANDOMNESS_VALUES.get(value)
            if flag is None:
                report.problem("непонятное значение случайности: «%s»" % cells[1])
                continue
            # В книге два отдельных «нет»: «Нет случайности» и «Полная
            # детерминированность». В опроснике им соответствует ОДИН вариант,
            # и называется он «Нет случайности (полная детерминированность)» —
            # то есть ровно вторая строка. Её и берём: это не «шаг посильнее»,
            # а точное совпадение с формулировкой, которую видел автор.
            key = "false" if flag is False else "true"
            exact = (value == "полная детерминированность")
            old = out[component][param].get(key)
            if old is None or exact or old.get("_weak"):
                entry["_weak"] = not exact and flag is False
                out[component][param][key] = entry
            continue

        if param == "adaptation":
            flag = YES_NO.get(value)
            if flag is None:
                report.problem("непонятное значение адаптации: «%s»" % cells[1])
                continue
            out[component][param]["true" if flag else "false"] = entry
            continue

        allowed = _survey_values(param)
        if allowed is not None and value not in allowed:
            # Таймер считается не по вариантам опросника, а по числовым
            # корзинам («1–4 игрока», «9+ игроков»). Это законный второй способ
            # задать поправку, а не опечатка, — распознаём его явно.
            bucket = parse_bucket(param, value)
            if bucket:
                out[component][param][bucket] = entry
                continue
            report.problem("значение «%s» (параметр %s, компонент %s) не совпало "
                           "ни с одним вариантом опросника"
                           % (cells[1], param, component))
            continue

        out[component][param][value] = entry
        seen_values.setdefault((component, param), set()).add(value)

    # Пропущенные варианты опросника — это молчаливый ноль в расчёте.
    for (component, param), values in sorted(seen_values.items()):
        allowed = _survey_values(param)
        if allowed is None:
            continue
        missing = allowed - values
        if missing:
            report.note("у «%s» / %s нет шага для: %s (примем 0)"
                        % (component, param, ", ".join(sorted(missing))))
    return out


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        return 1

    from openpyxl import load_workbook

    path = sys.argv[1]
    wb = load_workbook(path, data_only=True)
    report = Report()

    print("Читаю: %s" % path)
    base = read_base(wb, report)
    materials = read_materials(wb, report)
    steps = read_steps(wb, report)

    print("\nБазовые диапазоны: %d компонентов" % len(base))
    for key, row in base.items():
        print("  %-14s %d-%d (не ниже %d)%s" % (
            key, row["base_min"], row["base_max"], row["floor"],
            "  на игрока" if row["per_player"] else ""))
    print("\nМатериалы: %d компонентов" % len(materials))
    print("Поправки: %d компонентов" % len(steps))
    for key, params in steps.items():
        print("  %-14s параметров %d, значений %d"
              % (key, len(params), sum(len(v) for v in params.values())))

    print()
    ok = report.show()

    # Служебные пометки разбора наружу не выносим: в библиотеке должно лежать
    # только то, что читает расчёт.
    for params_of in steps.values():
        for table in params_of.values():
            for entry in table.values():
                entry.pop("_weak", None)

    data = {
        "version": 1,
        "source": Path(path).name,
        "comment": ("Собрано tools/build_components.py из книги Excel. "
                    "Правки вносите в книгу и пересобирайте, а не здесь: "
                    "иначе расчёт и таблица разойдутся."),
        "base": base,
        "materials": materials,
        "steps": steps,
    }
    OUT_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2),
                        encoding="utf-8")
    print("\nЗаписано: %s" % OUT_FILE)
    return 0 if ok else 2


if __name__ == "__main__":
    sys.exit(main())
